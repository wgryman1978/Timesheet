import PySimpleGUI as sg
import pandas as pd
import openpyxl as op

def weekly_wage():

    filename = 'Payroll.xlsx'

    wb = op.load_workbook(filename, data_only=True)

    ws = wb.worksheets[0]

    ws_tables = []


    layout = [
        [sg.Text('Employee Name', size=(15,1), font=('ariel', 16)), sg.InputText(key='Name', font=('ariel',16))],
        [sg.Text('Week Ending', size=(15,1), font=('ariel',16)), sg.InputText(key='Date', font=('ariel',16))],
        [sg.Submit(font=('ariel', 16))]
    ]

    window = sg.Window('Report Weekly Gross Wage', layout)#, size=(450,100), element_justification='c')

    # Event Loop

    def search():     #This is where I start to go wrong

        while True:
            event, values = window.read()
            key1 = window[key]('-Name-')       # this gives a callable object to the key
            key2 = window[key]('-Date-')         # another callable object
            for row in ws.iter_rows():
                if row[0].value == key1:   # filter on first column with value 16
                    if row[0].value == key2:
                        return ws.cell(row=cell.row, column=49).value as key3
                    else:
                        return '$0.00'as key3
                else:
                    return '$0.00' as key3
 
    def search_window():      #this is the window to display the value

        #key3 = search()
        
        layout = [
            [sg.Text('The total Gross Weekly Wage is:', font=('ariel', 19))],
            #[sg.Input(key='Key3', font=('ariel', 16), element_justification='c', background_color='Red')]
        ]

        window_1 = sg.Window('Wages Calculator', layout, size=(450,100), element_justification='c')
           
        while True:
            event, value = window.read()
            if event == sg.WIN_CLOSED:
                break
            
        window_1.close()
        
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED:
            break
        if event == 'Submit':
            search_window()

    window.close()
    
weekly_wage()
