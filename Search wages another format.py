
import PySimpleGUI as sg
import pandas as pd
import openpyxl as op

def weekly_wage():

    filename = 'Payroll.xlsx'

    wb = op.load_workbook(filename, data_only=True)

    ws = wb.worksheets[0]

    ws_tables = []


    layout = [
        [sg.Text('Employee Name', size=(15,1), font=('ariel', 16)), sg.InputText(key='-Name-', font=('ariel',16))],
        [sg.Text('Week Ending', size=(15,1), font=('ariel',16)), sg.InputText(key='-Date-', font=('ariel',16))],
        [sg.Submit(font=('ariel', 16))],
        [sg.Text('The Weekly Wage is:', font=('ariel', 16))],
        [sg.Output(size=(10, 1))]
    ]

    window = sg.Window('Gross Wages Search', layout, size=(450,250))

    # Event Loop
            
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED:
            break
        if event == 'Submit':
            key1 = window['-Name-']      # this gives a callable object to the key
            key2 = window['-Date-']         # another callable object
            for row in ws.iter_rows():
                if row[1].value == key1 and row[2].value == key2:   # filter on first column with value 16
                    ws.cell(row=cell.row, column=49).value
                    print(value, font=('ariel',16))
                    break
                else:
                    print('Check Your Search Parameters and Try Again', font=('ariel', 16))
                    break

    window.close()
    
""""def search_window():      #this is the window to display the value

        layout = [
            [sg.Text('Gross Weekly Wage Is:', font=('ariel', 19))],
            [sg.Output('-Key3-', font=('ariel', 16))]#, element_justification='c', background_color='Red')]
        ]

        window_1 = sg.Window('Wages Calculator', layout, size=(450,100), element_justification='c')

           
        while True:
            event, value = window.read()
            if event == sg.WIN_CLOSED:
                break

        window_1.close()"""""

weekly_wage()
